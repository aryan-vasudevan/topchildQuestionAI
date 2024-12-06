import openpyxl
import os
import json
from openai import OpenAI
from dotenv import load_dotenv
from pydantic import BaseModel

# Initialize API
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
INSTRUCTIONS = os.getenv("INSTRUCTIONS")
ASSISSTANT_ID = os.getenv("ASSISSTANT_ID")

client = OpenAI(api_key=OPENAI_API_KEY)

# Question model for structured outputs
class SampleQuestion(BaseModel):
    questionText: str
    answerChoices: list[str]
    correctAnswer: str

class NewQuestion(BaseModel):
    questionText: str
    answerChoices: list[str]
    correctAnswer: str
    explanation: str

# Get new question
thread = client.beta.threads.create()
def getNewQuestion(sampleQuestion, questionJSONList):

    message = client.beta.threads.messages.create(
        thread_id=thread.id,
        role="user",
        content=f"question - {sampleQuestion.questionText} \nanswer choices - {sampleQuestion.answerChoices} \ncorrect answer - {sampleQuestion.correctAnswer}"
    )

    run = client.beta.threads.runs.create_and_poll(
        thread_id=thread.id,
        assistant_id=ASSISSTANT_ID,
    )

    if run.status == "completed": 
        messages = list(client.beta.threads.messages.list(
            thread_id=thread.id,
    ))
    questionJSON = messages[0].content[0].text.value
    
    questionJSONList.append(questionJSON)
# Excel file
path = "questions.xlsx"
wb = openpyxl.load_workbook(path)

# Sample question sheet
sheet = wb.active

# New question sheet
numOfSheets = len(wb.sheetnames)
wb.create_sheet(f"Sheet {numOfSheets + 1}")

newSheet = wb[f"Sheet {numOfSheets + 1}"]

numRows = len([row for row in sheet if not all([cell.value == None for cell in row])])
numCols = 8

# Read questions
sampleQuestionList = []
for row in sheet.iter_rows(min_row=2, min_col=2, max_row=numRows, max_col=numCols):
    # Read specific cells
    questionText = str(row[0].value)
    answerChoices = [str(cell.value) for cell in row[1:6]]
    correctAnswer = str(row[6].value)

    # Keep questions in model form to keep it organized
    sampleQuestionList.append(SampleQuestion(questionText=questionText, answerChoices=answerChoices, correctAnswer=correctAnswer))

# Get a new question for each sample question
questionJSONList = []
for sampleQuestion in sampleQuestionList:
    getNewQuestion(sampleQuestion, questionJSONList)

newQuestionList = []
for questionJSON in questionJSONList:
    questionObj = json.loads(questionJSON)

    newQuestion = NewQuestion(questionText=questionObj["questionText"], answerChoices=questionObj["answerChoices"], correctAnswer=questionObj["correctAnswer"], explanation=questionObj["explanation"])
    newQuestionList.append(newQuestion)

for rowNumber, newQuestion in enumerate(newQuestionList):
    rowNumber += 1

    # Write question text
    newSheet.cell(row=rowNumber, column=1).value = newQuestion.questionText
    
    # Write answer choices
    for i in range(2, 7):
        newSheet.cell(row=rowNumber, column=i).value = newQuestion.answerChoices[i - 2]

    # Write correct answer
    newSheet.cell(row=rowNumber, column=7).value = newQuestion.correctAnswer

    # Write explanation
    newSheet.cell(row=rowNumber, column=8).value = newQuestion.explanation

wb.save(path)